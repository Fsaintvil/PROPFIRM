#!/usr/bin/env python3
"""Seed OnlineLearner with validated live trades from ReportHistory FTMO.

Remplace les seeds synthétiques par des VRAIS trades du compte FTMO Free Trial.
Ces 506 trades validés (XAUUSD, BTCUSD, ETHUSD, EURUSD, US500.cash) donnent
à l'OnlineLearner une base d'apprentissage réelle du marché.

Usage:
    python scripts/seed_from_reporthistory.py          # Import + sauvegarde
    python scripts/seed_from_reporthistory.py --dry    # Simulation sans écrire
"""

import argparse
import json
import logging
import sys
from collections import deque
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger("seed")

ACTIVE_SYMBOLS = {"XAUUSD", "BTCUSD", "ETHUSD", "EURUSD", "US500.cash"}
REPORT_FILE = "ReportHistory-1513621052.xlsx"
OL_STATE_FILE = Path("runtime/ol_state.json")
CAL_STATE_FILE = Path("runtime/ol_state.json")


def parse_reporthistory(xlsx_path: str) -> dict[str, list[dict]]:
    """Parse le fichier ReportHistory et extrait les trades par symbole."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl requis: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    trades: dict[str, list[dict]] = {s: [] for s in ACTIVE_SYMBOLS}

    for row in range(8, ws.max_row + 1):  # Data starts at row 8
        sym = ws.cell(row, 3).value
        if sym not in ACTIVE_SYMBOLS:
            continue

        typ = str(ws.cell(row, 4).value or "").lower()
        if typ not in ("buy", "sell"):
            continue

        profit_val = ws.cell(row, 13).value
        volume_val = ws.cell(row, 5).value
        if profit_val is None or volume_val is None or volume_val == 0:
            continue

        try:
            profit = float(profit_val)
        except (ValueError, TypeError):
            continue

        r = 1.0 if profit > 0 else -1.0
        regime = "BUY" if typ == "buy" else "SELL"
        trades[sym].append({"r": r, "regime": regime})

    return trades


def print_summary(trades: dict[str, list[dict]]):
    """Affiche le résumé des trades par symbole."""
    total = 0
    for sym in sorted(ACTIVE_SYMBOLS):
        t = trades.get(sym, [])
        if not t:
            continue
        wins = sum(1 for x in t if x["r"] > 0)
        total += len(t)
        logger.info(
            f"  {sym:12s}: {len(t):4d} trades, WR={wins / len(t) * 100:5.1f}%  (wins={wins}, losses={len(t) - wins})"
        )
    logger.info(f"  {'─' * 45}")
    logger.info(f"  {'TOTAL':12s}: {total:4d} trades validés")


def seed_online_learner(
    trades: dict[str, list[dict]],
    ol_path: Path,
    cal_path: Path,
    dry_run: bool = False,
):
    """Remplace les données seed de l'OnlineLearner par les trades validés.

    Principe:
      1. Charge l'état actuel de l'OnlineLearner
      2. Remplace l'historique de chaque symbole actif par les vrais trades
      3. Garde les autres symboles inchangés
      4. Sauvegarde l'état mis à jour
    """
    if not ol_path.exists():
        logger.warning(f"Fichier {ol_path} introuvable — création d'un nouvel état")
        state = {"window": 200, "history": {}, "adapted_params": {}}
    else:
        with open(ol_path) as f:
            state = json.load(f)

    window = state.get("window", 200)
    history = state.get("history", {})
    old_total = sum(len(v) for v in history.values())

    # Remplacer les symboles actifs par les vrais trades
    new_total = 0
    for sym in ACTIVE_SYMBOLS:
        sym_trades = trades.get(sym, [])
        if not sym_trades:
            logger.info(
                f"  {sym}: aucun trade validé — conservation des données existantes ({len(history.get(sym, []))} trades)"
            )
            new_total += len(history.get(sym, []))
            continue

        # Garder les 200 plus récents (limite de la fenêtre)
        if len(sym_trades) > window:
            sym_trades = sym_trades[-window:]
            logger.info(f"  {sym}: tronqué à {window} trades (fenêtre max)")

        history[sym] = sym_trades
        new_total += len(sym_trades)

    state["history"] = history

    # Recalculer adapted_params pour chaque symbole actif
    # (réinitialiser pour que le recalibrage se fasse au prochain trade live)
    adapted = state.get("adapted_params", {})
    for sym in ACTIVE_SYMBOLS:
        if sym in adapted:
            del adapted[sym]
            logger.info(f"  {sym}: adapted_params réinitialisés (seront recalculés au prochain trade)")

    if dry_run:
        logger.info(f"\n🔷 DRY RUN — Écriture simulée")
        logger.info(f"  {old_total} → {new_total} trades ({new_total - old_total:+d})")
        return

    # Sauvegarder online_learner_state.json
    with open(ol_path, "w") as f:
        json.dump(state, f, indent=2, default=str)
    logger.info(f"\n✅ {ol_path} mis à jour: {old_total} → {new_total} trades")

    # Mettre à jour calibration_state.json si présent
    if cal_path.exists():
        try:
            with open(cal_path) as f:
                cal_state = json.load(f)
            cal_state["online_history"] = history
            # Nettoyer adapted_params dans la calibration aussi
            if "adapted_params" in cal_state:
                for sym in ACTIVE_SYMBOLS:
                    cal_state["adapted_params"].pop(sym, None)
            with open(cal_path, "w") as f:
                json.dump(cal_state, f, indent=2, default=str)
            logger.info(f"✅ {cal_path} synchronisé ({new_total} trades)")
        except Exception as e:
            logger.warning(f"⚠️  Échec synchro calibration: {e}")


def main():
    parser = argparse.ArgumentParser(description="Seed OnlineLearner avec les trades validés ReportHistory")
    parser.add_argument(
        "--dry",
        action="store_true",
        help="Simulation sans écrire les fichiers",
    )
    parser.add_argument(
        "--file",
        default=REPORT_FILE,
        help=f"Chemin du fichier ReportHistory (défaut: {REPORT_FILE})",
    )
    args = parser.parse_args()

    report_path = Path(args.file)
    if not report_path.exists():
        logger.error(f"Fichier introuvable: {report_path}")
        sys.exit(1)

    logger.info("📊 Import des trades validés ReportHistory → OnlineLearner")
    logger.info(f"  Source: {report_path}")
    logger.info(f"  Symboles actifs: {', '.join(sorted(ACTIVE_SYMBOLS))}")
    logger.info("")

    trades = parse_reporthistory(str(report_path))
    print_summary(trades)

    total = sum(len(v) for v in trades.values())
    if total == 0:
        logger.error("Aucun trade trouvé pour les symboles actifs")
        sys.exit(1)

    logger.info("")
    logger.info("🚀 Seed OnlineLearner..." if not args.dry else "🔷 Simulation...")
    seed_online_learner(trades, OL_STATE_FILE, CAL_STATE_FILE, dry_run=args.dry)

    if not args.dry:
        logger.info("")
        logger.info("✅ Import terminé avec succès !")
        logger.info("  Les seeds synthétiques ont été remplacés par les vrais trades.")
        logger.info("  Les adapted_params seront recalculés au prochain trade live (FIX #1 actif).")


if __name__ == "__main__":
    main()
