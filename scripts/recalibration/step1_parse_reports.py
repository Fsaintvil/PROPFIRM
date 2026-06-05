"""Step 1: Parse ReportHistory Excel files into clean trade lists"""
import pickle
from collections import defaultdict

import openpyxl


def parse_report(path, account_label):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Sheet1']

    header_row = None
    col_map = {}
    for r, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [str(v) if v is not None else '' for v in row]
        joined = '|'.join(vals)
        if 'Profit' in joined and 'Symbole' in joined:
            header_row = r
            for i, h in enumerate(vals):
                h_lower = h.lower().strip()
                if 'symbole' in h_lower or 'symbol' in h_lower:
                    col_map['symbol'] = i
                elif 'type' in h_lower:
                    col_map['type'] = i
                elif 'volume' in h_lower:
                    col_map['volume'] = i
                elif 'profit' in h_lower:
                    col_map['profit'] = i
                elif h == 'Heure' and 'position' not in h.lower():
                    # First 'Heure' col is open time
                    if 'time_open' not in col_map:
                        col_map['time_open'] = i
                elif 'position' in h_lower:
                    col_map['position'] = i
            break

    print(f"Column map: {col_map}")

    trades = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = [v if v is not None else '' for v in row]
        try:
            profit_str = str(vals[col_map['profit']]).strip()
            if not profit_str or profit_str == '0':
                continue
            profit = float(profit_str.replace(',', '.'))
            if abs(profit) > 100000:
                continue  # corrupted

            symbol = vals[col_map['symbol']]
            if not symbol:
                continue

            trade_type = str(vals[col_map['type']]).lower().strip()
            if trade_type in ('buy', 'sell'):
                direction = trade_type.upper()
            else:
                continue

            time_open = vals[col_map['time_open']]
            time_str = time_open.strftime('%Y-%m-%d %H:%M:%S') if hasattr(time_open, 'strftime') else str(time_open)

            volume = float(vals[col_map['volume']]) if vals[col_map['volume']] else 0

            trades.append({
                'symbol': symbol,
                'direction': direction,
                'profit': profit,
                'volume': volume,
                'time_open': time_str[:19],
                'time_close': '',
                'won': profit > 0,
            })
        except (ValueError, TypeError, IndexError, KeyError):
            pass

    wb.close()

    # Deduplicate by position ID (last win/loss record)
    print(f"{account_label}: {len(trades)} raw trades")

    # Keep only one record per position - the closing one
    wins = sum(1 for t in trades if t['won'])
    losses = len(trades) - wins
    total_pnl = sum(t['profit'] for t in trades)
    print(f"  Wins: {wins}, Losses: {losses}, WR: {wins/len(trades)*100:.1f}%")
    print(f"  Total PnL: {total_pnl:.2f}")

    by_sym = defaultdict(lambda: {'pnl': 0, 'wins': 0, 'losses': 0})
    for t in trades:
        s = t['symbol']
        by_sym[s]['pnl'] += t['profit']
        if t['won']:
            by_sym[s]['wins'] += 1
        else:
            by_sym[s]['losses'] += 1
    for s, d in sorted(by_sym.items(), key=lambda x: x[1]['pnl']):
        tot = d['wins'] + d['losses']
        print(f"  {s}: {d['pnl']:+.2f} ({tot}t, WR={d['wins']/tot*100:.0f}%)")

    return trades

# Parse both reports
print("=== Report 1: 1513441721 (today) ===")
t1 = parse_report(
    "C:\\Users\\saint\\Documents\\MT5_FTMO_IA.7\\ReportHistory-1513441721.xlsx",
    "Account 1513441721"
)

print("\n=== Report 2: 1513284340 (May 18) ===")
t2 = parse_report(
    "C:\\Users\\saint\\Documents\\MT5_FTMO_IA.7\\ReportHistory-1513284340.xlsx",
    "Account 1513284340"
)

# Combine and save
combined = t1 + t2
print(f"\n=== COMBINED: {len(combined)} trades ===")
print(f"PnL: {sum(t['profit'] for t in combined):.2f}")

# Save for next steps
# Need to handle datetime objects for JSON
clean_trades = []
for t in combined:
    clean_trades.append({k: v for k, v in t.items()})

with open("runtime/historical_trades.pkl", "wb") as f:
    pickle.dump(clean_trades, f)
print(f"\nSaved {len(clean_trades)} trades to runtime/historical_trades.pkl")
