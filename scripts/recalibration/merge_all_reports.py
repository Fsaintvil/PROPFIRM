"""Parse TOUS les ReportHistory Excel avec le bon mapping de colonnes."""
import glob
import os
import pickle
from collections import Counter

import openpyxl


def parse_report(path, account_label, account_id):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Sheet1']

    header_row = None
    col_map = {}
    for r, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [str(v) if v is not None else '' for v in row]
        joined = '|'.join(vals)
        if 'Profit' in joined and 'Symbole' in joined:
            header_row = r
            for i, h in enumerate(vals):
                hl = h.lower().strip()
                if 'symbole' in hl or 'symbol' in hl:
                    col_map['symbol'] = i
                elif hl == 'type':
                    col_map['type'] = i
                elif 'volume' in hl:
                    col_map['volume'] = i
                elif 'profit' in hl:
                    col_map['profit'] = i
                elif hl == 'position':
                    col_map['position'] = i
                elif hl == 'heure' and 'position' not in hl:
                    if 'time_open' not in col_map:
                        col_map['time_open'] = i
                    else:
                        col_map['time_close'] = i
                elif hl == 'prix' and 'time_open' in col_map and 'close' not in col_map:
                    col_map['price_open'] = i
                elif hl == 'prix' and 'time_close' in col_map:
                    col_map['price_close'] = i
                elif 's / l' in hl or hl == 'sl':
                    col_map['sl'] = i
                elif 't / p' in hl or hl == 'tp':
                    col_map['tp'] = i
                elif 'commission' in hl:
                    col_map['commission'] = i
                elif 'echange' in hl or 'swap' in hl:
                    col_map['swap'] = i
            break

    if header_row is None:
        print("  No header found")
        wb.close()
        return []

    print(f"  Col map: {col_map}")

    trades = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = [v if v is not None else '' for v in row]
        try:
            profit_str = str(vals[col_map['profit']]).strip().replace(',', '.')
            if not profit_str or profit_str == '0':
                continue
            profit = float(profit_str)
            if abs(profit) > 100000:
                continue  # balance row

            symbol = vals[col_map['symbol']]
            if not symbol:
                continue

            trade_type = str(vals[col_map['type']]).lower().strip()
            if trade_type not in ('buy', 'sell'):
                continue

            time_open = vals[col_map['time_open']]
            if hasattr(time_open, 'strftime'):
                time_str = time_open.strftime('%Y-%m-%d %H:%M:%S')
            else:
                time_str = str(time_open).replace('T', ' ')[:19]

            volume = float(vals[col_map['volume']])
            price_open = float(vals[col_map['price_open']]) if col_map.get('price_open') is not None else 0
            price_close = float(vals[col_map['price_close']]) if col_map.get('price_close') is not None else 0
            sl = float(vals[col_map['sl']]) if col_map.get('sl') is not None else 0
            tp = float(vals[col_map['tp']]) if col_map.get('tp') is not None else 0
            commission = float(vals[col_map['commission']]) if col_map.get('commission') is not None else 0
            swap = float(vals[col_map['swap']]) if col_map.get('swap') is not None else 0

            pos_id = int(vals[col_map['position']]) if col_map.get('position') is not None else 0

            trades.append({
                "account": account_label, "account_id": account_id,
                "time_open": time_str,
                "symbol": symbol, "direction": trade_type,
                "volume": volume, "price_open": price_open,
                "price_close": price_close, "sl": sl, "tp": tp,
                "commission": commission, "swap": swap,
                "profit": round(profit, 2),
                "position_id": pos_id, "won": profit > 0,
            })
        except (ValueError, TypeError, KeyError):
            continue

    wb.close()
    return trades

ACCOUNTS = {
    1512422497: "FTMO_100K", 1512442430: "FTMO_100K_2",
    1512493997: "FTMO_100K_3", 1512497509: "FTMO_50K",
    1512498604: "FTMO_200K", 1512568197: "FTMO_200K_BIG",
    1513284340: "FTMO_200K_FAIL", 1513441721: "FTMO_200K_ACTIVE",
}

all_trades = []
for fpath in sorted(glob.glob("ReportHistory*.xlsx")):
    fname = os.path.basename(fpath)
    size = os.path.getsize(fpath)
    print(f"\n{fname} ({size//1024}KB)")

    acct_id = None
    for aid in ACCOUNTS:
        if str(aid) in fname:
            acct_id = aid
            break

    trades = parse_report(fpath, ACCOUNTS.get(acct_id or 0, fname), acct_id or 0)
    print(f"  -> {len(trades)} trades")
    all_trades.extend(trades)

print(f"\n{'='*60}")
print(f"TOTAL: {len(all_trades)} trades")

by_acct = Counter(t["account"] for t in all_trades)
print("\nPer account:")
for acct, count in sorted(by_acct.items(), key=lambda x: -x[1]):
    tt = [t for t in all_trades if t["account"] == acct]
    pnl = sum(t["profit"] for t in tt)
    won = sum(1 for t in tt if t["won"])
    print(f"  {acct:20s}: {count:5d}t, PnL={pnl:+.0f}, WR={won/count*100:.1f}%")

# Exclude failed account
print("\nExcluding FTMO_200K_FAIL...")
clean = [t for t in all_trades if t["account_id"] != 1513284340]
print(f"  -> {len(clean)} trades")

# Symbol breakdown
print("\nSymbol breakdown (clean):")
by_sym = Counter(t["symbol"] for t in clean)
for sym, count in sorted(by_sym.items(), key=lambda x: -x[1]):
    tt = [t for t in clean if t["symbol"] == sym]
    pnl = sum(t["profit"] for t in tt)
    won = sum(1 for t in tt if t["won"])
    print(f"  {sym:12s}: {count:5d}t, PnL={pnl:+8.0f}, WR={won/count*100:.1f}%")

times = sorted([t["time_open"] for t in clean])
print(f"\nDate range: {times[0][:10]} to {times[-1][:10]}")

with open("runtime/all_trades_full.pkl", "wb") as f:
    pickle.dump(all_trades, f)
with open("runtime/all_trades_clean.pkl", "wb") as f:
    pickle.dump(clean, f)
print(f"\nSaved runtime/all_trades_clean.pkl ({len(clean)} trades)")
