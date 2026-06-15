#!/usr/bin/env python3
"""Parse ReportHistory Excel files → seed CSV pour OnlineLearner.
Usage:
    python scripts/seed_from_excel.py
Output: runtime/online_learner_seed.csv
"""
import csv
import logging
import sys
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger("seed")

BASE = Path(__file__).resolve().parent.parent
EXCEL_FILES = [
    BASE / "ReportHistory-1513441721.xlsx",  # 2281 trades, Mai-Juin 2026
    BASE / "ReportHistory-1513621052.xlsx",  # 47 trades, Challenge en cours
]
OUTPUT = BASE / "runtime" / "online_learner_seed.csv"

# Mapping symbole → direction pour normaliser
DIR_MAP = {"buy": "BUY", "sell": "SELL", "Buy": "BUY", "Sell": "SELL"}

# R Multiple estimé : on utilise profit / (entry - sl) comme proxy du R
# Mais ces fichiers n'ont pas le SL de toutes les stratégies.
# On calcule r_multiple = profit / (volume * tick_value_estimate)
# Fallback: r_multiple = profit / max(abs(profit), 1) → signé ±1

def parse_excel(filepath: Path) -> list[dict]:
    """Returns list of dicts with keys: symbol, direction, volume, profit, r_multiple, timestamp."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Find header row with column names
    col_map = {}
    for r in range(1, 15):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str):
                vc = v.strip().lower()
                if vc in ("symbole", "symbol"):
                    col_map["symbol"] = c
                elif vc == "type":
                    col_map["direction"] = c
                elif vc == "volume":
                    col_map["volume"] = c
                elif vc in ("profit", "profit perte"):
                    col_map["profit"] = c
                elif vc in ("heure", "time", "position", "position id"):
                    col_map["timestamp"] = c

    if "symbol" not in col_map or "profit" not in col_map:
        logger.warning(f"{filepath.name}: colonnes symbol/profit non trouvées, ignoré")
        wb.close()
        return []

    logger.info(f"{filepath.name}: colonnes trouvées {col_map}")

    trades = []
    header_row = list(col_map.values())[0]  # approximate
    # Find actual header row
    actual_header = 7  # MT5 exports start at row 7 typically
    for r in range(1, 15):
        v = ws.cell(r, col_map["symbol"]).value
        if v and isinstance(v, str) and v.strip() in ("USDCAD", "EURUSD", "GBPUSD"):
            actual_header = r
            break
    # Header row is one before first data
    header_found = max(1, actual_header - 1)

    for r in range(header_found + 1, ws.max_row + 1):
        # Volume check
        vol_raw = ws.cell(r, col_map["volume"]).value if "volume" in col_map else "0"
        if vol_raw is None:
            continue
        try:
            vol = float(str(vol_raw).replace(",", "."))
        except (ValueError, TypeError):
            continue
        if vol <= 0:
            continue

        # Profit
        prof_raw = ws.cell(r, col_map["profit"]).value
        if prof_raw is None:
            continue
        try:
            profit = float(str(prof_raw).replace(",", ".").replace(" ", ""))
        except (ValueError, TypeError):
            continue

        # Symbol
        sym = str(ws.cell(r, col_map["symbol"]).value or "").strip()
        if not sym:
            continue

        # Direction
        direction = "?"
        if "direction" in col_map:
            d = str(ws.cell(r, col_map["direction"]).value or "").strip().lower()
            direction = DIR_MAP.get(d, d)

        # Timestamp
        ts = str(ws.cell(r, col_map.get("timestamp", 1)).value or "")

        # Estimate r_multiple: profit / (volume * 10) is a rough proxy
        # Standard forex: 1 lot = $10 per pip, so profit / (vol * 10 * pips)
        # Simplified: just normalize by absolute profit
        abs_profit = abs(profit)
        r_multiple = round(profit / abs_profit, 2) if abs_profit > 0 else 0
        # More refined: estimate risk per unit
        # For now, use a normalized R: profit / max(abs(profit), 0.01)
        # This gives ±1 for most trades, which is fine for the learner

        trades.append({
            "symbol": sym,
            "direction": direction,
            "volume": vol,
            "profit": profit,
            "r_multiple": r_multiple,
            "timestamp": ts[:19] if ts else "",
        })

    wb.close()
    logger.info(f"{filepath.name}: {len(trades)} trades extraits")
    return trades


def main():
    all_trades = []
    for f in EXCEL_FILES:
        if not f.exists():
            logger.warning(f"{f.name} non trouvé, ignoré")
            continue
        all_trades.extend(parse_excel(f))

    if not all_trades:
        logger.error("Aucun trade extrait — vérifie les fichiers Excel")
        sys.exit(1)

    # Écriture CSV
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["symbol", "direction", "volume", "profit", "r_multiple", "timestamp"]
    with open(OUTPUT, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_trades)

    logger.info(f"✅ {len(all_trades)} trades écrits dans {OUTPUT}")

    # Stats
    symbols = set(t["symbol"] for t in all_trades)
    wins = sum(1 for t in all_trades if t["profit"] > 0)
    losses = sum(1 for t in all_trades if t["profit"] < 0)
    logger.info(f"   Symboles: {len(symbols)} — {sorted(symbols)}")
    logger.info(f"   Wins={wins} Losses={losses} WR={wins/(wins+losses)*100:.1f}%")
    logger.info(f"   PnL total: ${sum(t['profit'] for t in all_trades):+.2f}")


if __name__ == "__main__":
    main()
