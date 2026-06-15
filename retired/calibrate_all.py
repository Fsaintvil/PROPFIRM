"""
calibrate_all.py — Importe les 951 trades historiques pour caler :
  - SQLite journal (stats/features)
  - OnlineLearner (seuils adaptatifs par symbole)
  - MetaLearner (poids dynamiques par régime)
  - Sauvegarde calibration → runtime/calibration_state.pkl

Usage: python calibrate_all.py
Puis redémarrer main.py pour qu'il charge la calibration.
"""

import logging
import os
import sys
from datetime import datetime, timedelta

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import MetaTrader5 as mt5

from engine_simple.adaptive_intelligence import MarketRegime, OnlineLearner
from engine_simple.meta_learner import MetaLearner

try:
    from engine_simple.ml_ensemble import MLEnsemble
    _ML_AVAILABLE = True
except ImportError:
    MLEnsemble = None
    _ML_AVAILABLE = False
import contextlib

from engine_simple.dl_ensemble import DLEnsemble
from engine_simple.trade_journal import TradeJournal

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger("calibrate")

for mod in ["ml_features", "ml_ensemble", "dl_ensemble", "signals", "adaptive", "ftmo_protector"]:
    logging.getLogger(mod).setLevel(logging.WARNING)

ROBOT_SYMBOLS = {
    "EURUSD", "GBPUSD", "USDJPY", "USDCAD", "USDCHF", "AUDUSD",
    "NZDUSD", "EURJPY", "GBPJPY", "XAUUSD", "ETHUSD", "USOIL.cash",
}
TF_ORDER = {1: "M1", 5: "M5", 15: "M15", 30: "M30", 60: "H1", 240: "H4", 1440: "D1"}

def main():
    # ─── 1. Parse ReportHistory Excel ───
    path = r"C:\Users\saint\Documents\MT5_FTMO_IA.7\ReportHistory-1513441721.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Sheet1']

    header_row = None
    for r, row in enumerate(ws.iter_rows(values_only=True), 1):
        vals = [str(v) if v is not None else '' for v in row]
        if 'Profit' in '|'.join(vals) and 'Symbole' in '|'.join(vals):
            header_row = r
            break

    raw_trades = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = [v if v is not None else '' for v in row]
        try:
            symbol = str(vals[2]).strip()
            if symbol not in ROBOT_SYMBOLS:
                continue
            trade_type = str(vals[3]).strip().lower()
            if trade_type not in ("buy", "sell"):
                continue
            profit_str = str(vals[12]).strip().replace(',', '.')
            if not profit_str:
                continue
            profit = float(profit_str)
            if abs(profit) > 1000:
                continue

            time_str = str(vals[0]).strip()
            try:
                dt = datetime.strptime(time_str, "%Y.%m.%d %H:%M:%S")
            except ValueError:
                dt = datetime.now()

            entry = float(vals[5]) if vals[5] != '' else 0
            sl = float(vals[6]) if vals[6] != '' else 0
            tp = float(vals[7]) if vals[7] != '' else 0
            lot = float(vals[4]) if vals[4] != '' else 0

            raw_trades.append({
                'symbol': symbol, 'direction': trade_type.upper(),
                'entry': entry, 'sl': sl, 'tp': tp, 'lot': lot,
                'profit': profit, 'won': profit > 0,
                'time': dt,
            })
        except (ValueError, TypeError, IndexError):
            pass
    wb.close()

    logger.info(f"Parsed {len(raw_trades)} valid trades from ReportHistory")
    raw_trades.sort(key=lambda t: t['time'])

    # ─── 2. MT5 Connect ───
    if not mt5.initialize():
        logger.error("MT5 init failed!")
        return
    logger.info("MT5 connected")

    # ─── 3. Init models ───
    regime_detector = MarketRegime()
    learner = OnlineLearner(window=50)
    meta = MetaLearner(recalibration_freq=50)
    ml_ensemble = MLEnsemble() if _ML_AVAILABLE else None
    dl_ensemble = DLEnsemble()
    journal = TradeJournal()

    # ─── 4. Group trades by symbol+hour ───
    from collections import OrderedDict
    buckets = OrderedDict()
    for t in raw_trades:
        key = (t['symbol'], t['time'].strftime("%Y-%m-%d %H"))
        buckets.setdefault(key, []).append(t)
    bucket_list = list(buckets.items())
    logger.info(f"Processing {len(bucket_list)} unique symbol/hour groups (total {len(raw_trades)} trades)")

    inserted, meta_fed, learner_fed = 0, 0, 0
    rates_cache = {}

    for bidx, ((symbol, hour_str), trades_in_bucket) in enumerate(bucket_list):
        bucket_time = datetime.strptime(hour_str, "%Y-%m-%d %H")
        cache_key = f"{symbol}_{hour_str}"

        # Fetch H1 rates
        if cache_key not in rates_cache:
            dt_from = bucket_time - timedelta(hours=400)
            h1_data = mt5.copy_rates_from(symbol, mt5.TIMEFRAME_H1, dt_from, 400)
            if h1_data is None or len(h1_data) < 50:
                # retry with wider range
                dt_from = bucket_time - timedelta(hours=600)
                h1_data = mt5.copy_rates_from(symbol, mt5.TIMEFRAME_H1, dt_from, 500)
            if h1_data is None or len(h1_data) < 50:
                logger.warning(f"  [{bidx}/{len(bucket_list)}] {symbol}@{hour_str}: "
                              f"no H1 data, skip {len(trades_in_bucket)} trades")
                continue
            h1_list = [tuple(r) for r in h1_data] if hasattr(h1_data, 'dtype') else list(h1_data)
            rates_cache[cache_key] = h1_list
        else:
            h1_list = rates_cache[cache_key]

        # Rates before bucket_time
        cutoff_ts = bucket_time.timestamp()
        active = [r for r in h1_list if r[0] <= cutoff_ts]
        if len(active) < 50:
            continue

        # Regime
        regime, meta_info = regime_detector.detect(active)

        # ML + DL predictions
        rd = {"H1": active}
        ml_result = None
        if ml_ensemble is not None:
            with contextlib.suppress(Exception):
                ml_result = ml_ensemble.predict(symbol, rd)
        dl_result = None
        if dl_ensemble.available:
            with contextlib.suppress(Exception):
                dl_result = dl_ensemble.predict(symbol, rd)

        ml_action = ml_result.get("action", "HOLD") if ml_result else None
        dl_action = dl_result.get("action", "HOLD") if dl_result else None

        for trade in trades_in_bucket:
            try:
                pos_dir = trade['direction']
                symbol = trade['symbol']

                # Predictions dict
                preds = {"MOM20x3": pos_dir}
                if ml_action and ml_action != "HOLD":
                    preds["RF"] = ml_action
                    preds["XGB"] = ml_action
                    preds["LGBM"] = ml_action
                if dl_action and dl_action != "HOLD":
                    preds["DL_LSTM"] = dl_action

                # Outcomes
                pc = trade['won']
                outcomes = {m: (a == pos_dir) if pc else (a != pos_dir) for m, a in preds.items()}

                # Feed MetaLearner
                meta.record_trade(symbol, regime, outcomes)
                meta_fed += 1

                # R-multiple via MT5 order_calc_profit
                ot = 0 if pos_dir == "BUY" else 1
                r1_usd = abs(mt5.order_calc_profit(ot, symbol, trade['lot'], trade['entry'], trade['sl']) or 0)
                if r1_usd > 0:
                    r_mul = trade['profit'] / r1_usd
                    learner.record_trade(symbol, r_mul, regime)
                    learner_fed += 1

                # SQLite
                journal.record({
                    "symbol": symbol, "direction": pos_dir,
                    "entry": trade['entry'], "sl": trade['sl'], "tp": trade['tp'],
                    "lot": trade['lot'], "profit": trade['profit'],
                    "time_open": trade['time'].isoformat(),
                    "time_close": trade['time'].isoformat(),
                    "reason": "history_calibrate"
                })
                inserted += 1
            except Exception as e:
                logger.debug(f"  Trade error: {e}")

        if (bidx + 1) % 15 == 0 or bidx == len(bucket_list) - 1:
            logger.info(f"  [{bidx+1}/{len(bucket_list)}] {inserted}/{len(raw_trades)} trades")

    # ─── 5. Save calibration state ───
    import joblib
    os.makedirs("runtime", exist_ok=True)

    cal_data = {
        "meta_regime_performance": dict(meta.regime_performance),
        "meta_trades_since_recal": meta.trades_since_recal,
        "meta_trackers": {},
    }
    for name, tracker in meta.trackers.items():
        cal_data["meta_trackers"][name] = {
            "regime_stats": dict(tracker.regime_stats),
            "symbol_stats": dict(tracker.symbol_stats),
            "global_stats": dict(tracker.global_stats),
        }

    ol_data = {}
    for sym, hist in learner.history.items():
        ol_data[sym] = list(hist)

    state = {
        "meta_calibration": cal_data,
        "online_history": ol_data,
        "calibrated_at": datetime.now().isoformat(),
        "total_trades": len(raw_trades),
    }
    joblib.dump(state, "runtime/calibration_state.pkl")
    logger.info("Saved calibration state to runtime/calibration_state.pkl")

    # ─── 6. Summary ───
    logger.info(f"\n{'='*55}")
    logger.info("CALIBRATION COMPLETE")
    logger.info(f"{'='*55}")
    logger.info(f"SQLite inserted: {inserted}")
    logger.info(f"MetaLearner fed: {meta_fed}")
    logger.info(f"OnlineLearner fed: {learner_fed}")
    logger.info("")

    for regime in ["TREND_UP", "TREND_DOWN", "RANGING", "HIGH_VOL", "LOW_VOL"]:
        wr = meta.get_regime_win_rate(regime)
        w = meta.get_weights(regime)
        top = sorted(w.items(), key=lambda x: -x[1])[:3]
        logger.info(f"  {regime:12s} WR={wr:.1%}  best: {', '.join(f'{m}={w_:.2f}' for m,w_ in top)}")

    logger.info("")
    for s in sorted(set(t['symbol'] for t in raw_trades)):
        sm = learner.get_summary(s)
        if sm:
            p = learner.get_params(s)
            logger.info(f"  {s:12s} {sm['trades']:4d}t WR={sm['wr']:.0%} "
                        f"expectancy={sm['expectancy']:+.2f}  "
                        f"thresh={p['thresh']} risk={p['risk_mult']:.2f}")

    mt5.shutdown()
    logger.info("\nDone! Redémarre main.py pour charger la calibration.")

if __name__ == "__main__":
    main()
